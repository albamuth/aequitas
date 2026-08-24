"""
track1_labour.py -- Track 1 of the median-adult lifestyle cost (bottom-up).

Computes the DOMESTIC human labour-hours embodied in US personal consumption,
per person / adult, using real EEIO data:

    hours = SUM_over_commodities[ PCE(commodity) x jobs_per_$M(commodity) ] x hours_per_job

- PCE(commodity)      : US personal consumption by commodity, PRODUCER values,
                        margins already allocated to trade/transport sectors.
                        BLS IO final-demand file, 2023, column 1 (verified = $18.82T).
- jobs_per_$M(commodity): total (direct + indirect) jobs supported per $1M of
                        final demand for that commodity. BLS Employment Requirements
                        Matrix 2023, DOMESTIC (import-adjusted) -> imports are Track 3.
                        = column-sum of the matrix (values are THOUSANDS of jobs) x 1000.
- hours_per_job       : jobs are a HEADCOUNT (incl. part-time + self-employed), not
                        FTE, so we convert to hours with an explicit, documented factor.

This is the rigorous replacement for the rejected top-down v1: no assumed labour
allocation (the ERM measures the whole supply chain) and no invented margins (the
BLS bridge already put spending in producer values).

DATA (recovered Wayback vintages; BLS withdrew the live tables 2026-02-06):
    data/erm_full/NOMINAL_DOMEMPREQ_2023.csv   (176x176 employment requirements)
    data/erm_full/SectorPlan2034.xlsx          (sector titles)
    data/io/NOMINAL_FDAGG.xlsx                 (final demand by commodity; col 1 = PCE)

Run:  python track1_labour.py
      python track1_labour.py --test
"""
from __future__ import annotations

import argparse
import csv
import os
from dataclasses import dataclass

import numpy as np
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
ERM_CSV = os.path.join(DATA, "erm_full", "NOMINAL_DOMEMPREQ_2023.csv")
SECTOR_XLSX = os.path.join(DATA, "erm_full", "SectorPlan2034.xlsx")
FDAGG_XLSX = os.path.join(DATA, "io", "NOMINAL_FDAGG.xlsx")

PCE_COLUMN = 1          # FDAGG column 1 = Personal Consumption Expenditures (verified)
FD_YEAR = "2023"


@dataclass(frozen=True)
class Params:
    # Jobs in the ERM are a headcount incl. part-time & self-employed (not FTE).
    # Average annual hours per job across ALL US jobs. ~1,750 is a defensible mid
    # (full-time ~1,800 OECD; part-time pulls the all-job average down). Sensitivity
    # reported below.
    hours_per_job: float = 1_750.0
    population: float = 335e6        # US resident population 2023 (~334.9M)
    adults_18plus: float = 258e6     # US 18+ (~77% of population)
    consumer_units: float = 134.556e6  # CE 2023 number of consumer units (thousands*1e3)
    # Consumption is less unequal than income; median consumer-unit spending is
    # below the mean. ~0.83 is a documented placeholder (consumption Gini ~0.3),
    # to be refined from CE microdata. Flagged.
    median_over_mean: float = 0.83


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_erm(path=ERM_CSV):
    """Return (sector_ids, jobs_per_M): total jobs per $1M final demand by commodity."""
    with open(path) as f:
        r = csv.reader(f)
        header = next(r)
        rows = [row for row in r]
    col_ids = [int(float(x)) for x in header[1:]]          # commodity j on columns
    M = np.array([[float(x) for x in row[1:]] for row in rows])   # thousands jobs/$1M
    # total (direct+indirect) jobs per $1M FD for commodity j = column sum * 1000
    jobs_per_M = M.sum(axis=0) * 1000.0
    return col_ids, jobs_per_M


def load_sector_titles(path=SECTOR_XLSX):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Stubs"]
    title = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            sn = int(row[0])
        except (TypeError, ValueError):
            continue
        title[sn] = str(row[3]) if row[3] else ""
    return title


def load_pce(path=FDAGG_XLSX, year=FD_YEAR, col=PCE_COLUMN):
    """Return {sector_id: PCE in $millions} for the given year (producer values)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[year]
    pce = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            continue
        try:
            sid = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        v = row[col]
        pce[sid] = float(v) if isinstance(v, (int, float)) else 0.0
    return pce


# ---------------------------------------------------------------------------
# Core computation
# ---------------------------------------------------------------------------

def compute(p: Params | None = None):
    p = p or Params()
    sector_ids, jobs_per_M = load_erm()
    pce = load_pce()
    jpm = {sid: jobs_per_M[i] for i, sid in enumerate(sector_ids)}

    # jobs supporting PCE of commodity j = PCE($M) * jobs_per_$1M
    rows = []
    total_jobs = 0.0
    for sid in sector_ids:
        pce_m = pce.get(sid, 0.0)          # $millions
        jobs = pce_m * jpm[sid]            # $M * (jobs per $1M) = jobs
        total_jobs += jobs
        rows.append((sid, pce_m, jpm[sid], jobs))

    total_hours = total_jobs * p.hours_per_job
    pce_total_M = sum(pce.values())

    per_capita = total_hours / p.population
    per_adult = total_hours / p.adults_18plus
    per_cu = total_hours / p.consumer_units
    per_adult_median = per_adult * p.median_over_mean

    return dict(
        params=p, rows=rows, total_jobs=total_jobs, total_hours=total_hours,
        pce_total_M=pce_total_M, per_capita=per_capita, per_adult=per_adult,
        per_cu=per_cu, per_adult_median=per_adult_median,
    )


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def report(p: Params | None = None):
    p = p or Params()
    res = compute(p)
    titles = load_sector_titles()
    W = 78
    print("=" * W)
    print("TRACK 1 -- domestic labour-hours embodied in US personal consumption (2023)")
    print("=" * W)
    print(f"PCE total (producer values) : ${res['pce_total_M']/1e6:6.2f} trillion  "
          f"(check: US PCE 2023 ~$18.8T)")
    print(f"Jobs supporting all PCE     : {res['total_jobs']/1e6:6.1f} million  "
          f"(check: ~0.68 x 160M US jobs ~ 109M)")
    print(f"hours/job (headcount)       : {p.hours_per_job:,.0f}")
    print(f"TOTAL PCE labour            : {res['total_hours']/1e9:6.1f} billion hours/yr")
    print("-" * W)
    print("Per-person embodied DOMESTIC labour (Track 1 only; imports = Track 3):")
    print(f"  per capita (all persons)  : {res['per_capita']:6.0f} h/yr")
    print(f"  per ADULT (18+)           : {res['per_adult']:6.0f} h/yr   <- 'support one adult'")
    print(f"  per consumer unit (h/hold): {res['per_cu']:6.0f} h/yr")
    print(f"  per MEDIAN adult (x{p.median_over_mean}) : {res['per_adult_median']:6.0f} h/yr"
          f"   <- headline (mean->median, flagged)")
    print("-" * W)
    # sensitivity on hours_per_job
    lo = compute(Params(hours_per_job=1650)); hi = compute(Params(hours_per_job=1850))
    print(f"sensitivity (hours/job 1650-1850): per-adult "
          f"{lo['per_adult']:.0f}-{hi['per_adult']:.0f} h/yr")
    print("-" * W)
    print("Top 12 commodities by embodied PCE labour-hours:")
    rows = sorted(res["rows"], key=lambda r: r[3], reverse=True)[:12]
    print(f"  {'commodity':<44}{'PCE $B':>9}{'jobs/$M':>9}{'B hrs':>8}")
    for sid, pce_m, jpm_v, jobs in rows:
        bhrs = jobs * p.hours_per_job / 1e9
        print(f"  {titles.get(sid,'')[:43]:<44}{pce_m/1e3:>9.0f}{jpm_v:>9.1f}{bhrs:>8.1f}")
    print("=" * W)
    print("NOTE: DOMESTIC labour only (ERM is import-adjusted). Housing services here")
    print("are owner-occupied-dwelling + rent (low labour); the CONSTRUCTION labour of")
    print("the housing stock is Track 2 (bill-of-materials). Imports = Track 3.")
    print("=" * W)
    return res


# ---------------------------------------------------------------------------
# Self-tests
# ---------------------------------------------------------------------------

def test_pce_total_matches_known():
    res = compute()
    tril = res["pce_total_M"] / 1e6
    assert 18.0 < tril < 19.5, f"PCE total ${tril:.2f}T should be ~$18.8T"
    print(f"[ok] PCE total ${tril:.2f}T matches US 2023 (~$18.8T)")


def test_pce_jobs_plausible():
    res = compute()
    m = res["total_jobs"] / 1e6
    assert 90 < m < 130, f"PCE-supported jobs {m:.0f}M should be ~109M"
    print(f"[ok] PCE-supported jobs {m:.0f}M (~0.68 x 160M US jobs)")


def test_negatives_are_minor():
    """A few commodities are legitimately negative in PCE (net used-goods purchases,
    the noncomparable-imports adjustment). Confirm the total is positive and those
    negatives are a small share, rather than wrongly forbidding them."""
    res = compute()
    neg = sum(r[3] for r in res["rows"] if r[3] < 0)
    assert res["total_jobs"] > 0
    assert abs(neg) < 0.05 * res["total_jobs"], (
        f"negative contributions {neg/1e6:.1f}M should be <5% of total "
        f"{res['total_jobs']/1e6:.0f}M")
    print(f"[ok] total positive; legit negatives (used goods/adjustments) "
          f"{neg/1e6:.1f}M jobs = {100*abs(neg)/res['total_jobs']:.1f}% (minor)")


def test_per_adult_range():
    res = compute()
    h = res["per_adult"]
    assert 300 < h < 1200, f"per-adult {h:.0f} h/yr implausible"
    print(f"[ok] per-adult domestic labour {h:.0f} h/yr in sane range")


def run_tests():
    test_pce_total_matches_known()
    test_pce_jobs_plausible()
    test_negatives_are_minor()
    test_per_adult_range()
    print("\nAll self-tests passed.")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--test", action="store_true")
    args = ap.parse_args()
    if args.test:
        run_tests()
        return 0
    report()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
